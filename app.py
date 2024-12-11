import os
from flask import Flask, jsonify, request
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

app=Flask(__name__)

UPLOAD_FOLDER = r'C:\Users\ADMIN\Desktop\student Management  uploading from CSV file\Upload'
ALLOWED_EXTENSIONS = {"csv","xlsx"}


db = SQLAlchemy()
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///students.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    age = db.Column(db.Integer, nullable=False)
    City = db.Column(db.String(20), nullable=False)

db.init_app(app)
with app.app_context():
    db.create_all()

@app.route('/add',methods=["POST"])
def add_students():
    data = request.get_json()
    new_student = Student(name=data['name'], age=data['age'], City=data['City'])
    db.session.add(new_student)
    db.session.commit()
    return "Done"

@app.route('/view',methods=['GET'])
def Get_detials():
    tasks = Student.query.all()
    task_list = [{'id':task.id,'name':task.name,'age':task.age,'City':task.City} for task in tasks ]
    return jsonify(task_list), 200

@app.route('/by_id/<int:id>',methods=['GET'])
def get_by_id(id):
    task = Student.query.get(id)
    if task:
        return jsonify({'id': task.id, 'name': task.name, 'age': task.age, 'City': task.City}), 200
    else:
        return jsonify({'error': 'Student not found'}), 404
    

@app.route('/update_student/<int:id>',methods=['PUT'])
def update_student(id):
    tasks = Student.query.get(id)
    if not tasks:
        return "Not Found"
    data = request.get_json(id)
    tasks.name = data.get('name', tasks.name)
    tasks.age = data.get('age', tasks.age)
    tasks.City = data.get('City', tasks.City)
    db.session.commit()
    return jsonify({"message": "Task updated successfully", "task": {"id": tasks.id, "name": tasks.name, "age": tasks.age,"City":tasks.City}}), 200
    
@app.route('/delete_students/<int:id>',methods=['DELETE'])
def delete_student(id):
    task = Student.query.get(id)
    if not task:
        return jsonify({"message": "Task not found"}), 404
    db.session.delete(task)
    db.session.commit()
    return jsonify({"message": "Task deleted successfully"}), 200

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/upload/excel', methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"message": "Error: No file found"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"message": "No file selected"}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

        workbook = load_workbook(file)
        sheet = workbook.active
        count = 0
        for row in sheet.iter_rows(min_row=2,values_only=True):
            count += 1
            student = Student(name=row[0],age=row[1],City=row[2])
            db.session.add(student)
            db.session.commit()
        return f"Sucessfully Added {count} data to the database"
    return ""



if __name__ == '__main__':
    app.run(debug=True)